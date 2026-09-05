# ============================================================
#   Импорт категорий фона (fon_type_N) из эксель-отчёта в price_checks.
#
#   Запуск:
#       python 03_import_fon_type.py "Образе_много_шаблонов.xlsx"
#
#   Логика: в экселе колонка "Исходное имя" = ваше реальное имя файла
#   (то же самое, что price_checks.filename). Колонка "Категория фона"
#   переносится в новое поле price_checks.fon_type по совпадению
#   имени файла. Если такого filename в базе нет — строка пропускается
#   и в конце выводится список пропущенных (для проверки — либо файл
#   ещё не был загружен, либо разошлись имена).
#
#   Можно запускать многократно на разных экспортах — просто
#   обновляет поле, не трогает остальные данные строки.
# ============================================================
import sys
import openpyxl
from supabase import create_client
import os

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

if not (SUPABASE_URL and SUPABASE_SERVICE_KEY):
    raise SystemExit("Задайте SUPABASE_URL и SUPABASE_SERVICE_KEY в переменных окружения.")

if len(sys.argv) < 2:
    raise SystemExit("Использование: python 03_import_fon_type.py <путь_к_xlsx> [имя_листа]")

xlsx_path = sys.argv[1]
sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb[sheet_name] if sheet_name else wb.active

header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
try:
    name_col = header.index("Исходное имя")
    fon_col = header.index("Категория фона")
except ValueError:
    raise SystemExit(
        f"Не нашёл колонки 'Исходное имя' / 'Категория фона' в листе. "
        f"Реальные колонки: {header}"
    )

# Тянем все известные filename -> id одним запросом, чтобы не дёргать
# базу по одной строке на каждую запись экселя.
all_rows = []
start = 0
page_size = 1000
while True:
    result = (supabase.table("price_checks")
              .select("id, filename")
              .range(start, start + page_size - 1)
              .execute())
    chunk = result.data
    all_rows.extend(chunk)
    if len(chunk) < page_size:
        break
    start += page_size

filename_to_id = {r["filename"]: r["id"] for r in all_rows}

updated = 0
not_found = []
for row in ws.iter_rows(min_row=2, values_only=True):
    fname = row[name_col]
    fon_type = row[fon_col]
    if not fname or not fon_type:
        continue
    row_id = filename_to_id.get(str(fname))
    if row_id is None:
        not_found.append(fname)
        continue
    supabase.table("price_checks").update({"fon_type": str(fon_type)}).eq("id", row_id).execute()
    updated += 1

print(f"Обновлено записей: {updated}")
if not_found:
    print(f"Не найдено в базе (пропущено): {len(not_found)}")
    for fn in not_found[:20]:
        print("  -", fn)
    if len(not_found) > 20:
        print(f"  ... и ещё {len(not_found) - 20}")
